"""
Flexible parser for the fuel-price source file.

The assessment ships a fuel-price file whose exact column layout is not
guaranteed. Real-world variants seen in this challenge include:

  * files WITH coordinates: name, address, city, state, latitude, longitude,
    price/retail price ...
  * files WITHOUT coordinates (the common "OPIS Truckstop" export):
    OPIS Truckstop ID, Truckstop Name, Address, City, State, Rack ID,
    Retail Price

This module normalises headers, maps a wide set of synonyms to a canonical
schema, and -- when coordinates are absent -- falls back to a US
state-centroid table (and, optionally, real geocoding driven by the import
command). The goal is that *any* reasonable fuel file just works.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator

from .us_centroids import STATE_CENTROIDS

# Canonical field -> accepted header synonyms (compared lower/trimmed).
COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "name": ("truckstop name", "station name", "name", "station", "site name"),
    "brand": ("brand", "company", "chain"),
    "address": ("address", "street", "address line 1", "addr"),
    "city": ("city", "town"),
    "state": ("state", "st", "state code"),
    "latitude": ("latitude", "lat"),
    "longitude": ("longitude", "lng", "lon", "long"),
    "price": (
        "retail price",
        "price per gallon",
        "price",
        "retail",
        "diesel price",
        "gasoline price",
        "gas price",
    ),
    "source_id": (
        "opis truckstop id",
        "truckstop id",
        "site id",
        "station id",
        "id",
        "rack id",
    ),
}


@dataclass
class ParsedStation:
    name: str
    state: str
    price_per_gallon: float
    brand: str | None = None
    address: str | None = None
    city: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    source_id: str | None = None
    is_approximate_location: bool = False


class FuelFileError(Exception):
    """Raised when the fuel file cannot be parsed at all."""


def _norm(header: str) -> str:
    return header.strip().lower().replace("_", " ")


def _build_header_map(headers: Iterable[str]) -> dict[str, int]:
    """Map canonical field -> column index using the synonym table."""
    index: dict[str, int] = {}
    for col_idx, raw in enumerate(headers):
        key = _norm(raw)
        for canonical, synonyms in COLUMN_SYNONYMS.items():
            if canonical in index:
                continue
            if key in synonyms:
                index[canonical] = col_idx
    return index


def _rows_from_csv(path: Path) -> tuple[list[str], Iterator[list[str]]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise FuelFileError("CSV file is empty.")
    return rows[0], iter(rows[1:])


def _rows_from_xlsx(path: Path) -> tuple[list[str], Iterator[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise FuelFileError(
            "openpyxl is required to read .xlsx files (pip install openpyxl)."
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [
        ["" if c is None else str(c) for c in row]
        for row in ws.iter_rows(values_only=True)
    ]
    if not rows:
        raise FuelFileError("XLSX sheet is empty.")
    return rows[0], iter(rows[1:])


def _to_float(value: str) -> float | None:
    if value is None:
        return None
    cleaned = str(value).strip().replace("$", "").replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_fuel_file(file_path: str | Path) -> list[ParsedStation]:
    """Parse a CSV/XLSX fuel file into a list of :class:`ParsedStation`.

    Rows missing a name, state, or a parseable price are skipped (logged by
    the caller). Rows missing coordinates fall back to the state centroid and
    are flagged ``is_approximate_location`` so the optimizer/notes can warn.
    """
    path = Path(file_path)
    if not path.exists():
        raise FuelFileError(f"Fuel file not found: {path}")

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        headers, body = _rows_from_xlsx(path)
    else:
        headers, body = _rows_from_csv(path)

    hmap = _build_header_map(headers)
    missing_required = [f for f in ("name", "state", "price") if f not in hmap]
    if missing_required:
        raise FuelFileError(
            "Could not locate required column(s) "
            f"{missing_required} in headers {headers!r}. "
            "Edit COLUMN_SYNONYMS in fuel_parser.py to add your file's names."
        )

    def cell(row: list[str], field: str) -> str | None:
        idx = hmap.get(field)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return val.strip() if isinstance(val, str) else val

    stations: list[ParsedStation] = []
    for row in body:
        if not row or not any(str(c).strip() for c in row):
            continue
        name = cell(row, "name")
        state = (cell(row, "state") or "").upper()[:8]
        price = _to_float(cell(row, "price"))
        if not name or not state or price is None:
            continue

        lat = _to_float(cell(row, "latitude"))
        lng = _to_float(cell(row, "longitude"))
        approximate = False
        if lat is None or lng is None:
            centroid = STATE_CENTROIDS.get(state)
            if centroid is None:
                # Unknown state and no coords -> cannot place it on a map.
                continue
            lat, lng = centroid
            approximate = True

        stations.append(
            ParsedStation(
                name=name,
                state=state,
                price_per_gallon=price,
                brand=cell(row, "brand") or None,
                address=cell(row, "address") or None,
                city=cell(row, "city") or None,
                latitude=lat,
                longitude=lng,
                source_id=cell(row, "source_id") or None,
                is_approximate_location=approximate,
            )
        )

    if not stations:
        raise FuelFileError(
            "No valid station rows parsed. Check the file has name/state/price."
        )
    return stations
